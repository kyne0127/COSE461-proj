#!/usr/bin/env python3
"""
scripts/make_feedback_csv.py

sequence_summary.json + manifest.jsonl을 합쳐 피드백 작성용 XLSX를 생성합니다.
객체별(종류) 탐지 개수 컬럼을 실제 레이블명으로 포함합니다.

로컬 실행:
    python scripts/make_feedback_csv.py \
        --summary  c:/tmp/det_sequence_vis/sequence_summary.json \
        --out      c:/tmp/det_sequence_vis/feedback.xlsx
"""

import argparse
import json
import collections
from pathlib import Path

import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from huggingface_hub import hf_hub_download

REPO_EVAL  = "kyne0127/vla-evaluation"
REPO_TRAIN = "kyne0127/ambres-training"

# 객체별 헤더 색상 (light tones)
OBJ_COLORS = {
    "cup":       "DDEEFF",   # 파란 계열
    "red box":   "FFE0E0",   # 빨간 계열
    "cube":      "E8F5E9",   # 초록 계열
    "box":       "FFF9C4",   # 노란 계열
}
DEFAULT_OBJ_COLOR = "F3E5F5"  # 연보라 (미지정 객체)

# 섹션별 헤더 색상
HEADER_COLORS = {
    "meta":       "CFE2F3",   # 기본 정보
    "gt":         "E8D5B7",   # GT 입력
    "gdino":      "D9EAD3",   # GDino 탐지 수
    "gdino_note": "B6D7A8",   # GDino 노트
    "molmo":      "FCE5CD",   # Molmo 탐지 수
    "molmo_note": "F9CB9C",   # Molmo 노트
    "flag":       "EAD1DC",   # 플래그
    "feedback":   "FFF2CC",   # 피드백
}


def load_manifest_hf(repo_id: str, filename: str):
    cached = hf_hub_download(repo_id, repo_type="dataset", filename=filename)
    entries = [json.loads(l) for l in open(cached, encoding="utf-8") if l.strip()]
    # 중복 id 제거 (첫 번째 유지)
    seen: set = set()
    deduped = []
    for e in entries:
        if e["id"] not in seen:
            seen.add(e["id"])
            deduped.append(e)
    return {e["id"]: e for e in deduped}


def flag(t0, c1, c2):
    if t0 > 0 and c1 == 0 and c2 == 0:
        return "disappear_both"
    if t0 > 0 and (c1 == 0 or c2 == 0):
        return "disappear_partial"
    if t0 == 0 and (c1 > 0 or c2 > 0):
        return "force_map"
    if t0 > 1 or c1 > 1 or c2 > 1:
        return "multi_detect"
    return ""


def hex_fill(hex_color: str) -> PatternFill:
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--summary", default="c:/tmp/det_sequence_vis/sequence_summary.json")
    parser.add_argument("--out",     default="c:/tmp/det_sequence_vis/feedback.xlsx")
    parser.add_argument("--dataset", default="eval", choices=["eval", "train"],
                        help="eval=vla-evaluation, train=ambres-training")
    args = parser.parse_args()

    rows     = json.load(open(args.summary, encoding="utf-8"))
    if args.dataset == "train":
        manifest = load_manifest_hf(REPO_TRAIN, "manifest_train.jsonl")
    else:
        manifest = load_manifest_hf(REPO_EVAL, "dataset/manifest.jsonl")

    # 데이터셋 전체 unique 객체 종류 (등장 순서 유지)
    all_objects: list[str] = []
    seen: set[str] = set()
    for r in rows:
        for obj in r["objects"]:
            if obj not in seen:
                all_objects.append(obj)
                seen.add(obj)

    # ── 행 구성 ────────────────────────────────────────────────────────────────
    data = []
    for r in rows:
        sid  = r["id"]
        meta = manifest.get(sid, {})
        tgt  = meta.get("target_label",      r["objects"][0] if r["objects"] else "")
        dst  = meta.get("destination_label",  r["objects"][1] if len(r["objects"]) > 1 else "")

        def g(det, frame, obj):
            return r.get(f"{det}_{frame}_{obj}", 0)

        row: dict = {
            "id":                sid,
            "scenario":          meta.get("scenario", ""),
            "task":              meta.get("task", ""),
            "objects":           ", ".join(r["objects"]),     # 종류 한눈에
            "target_label":      tgt,
            "destination_label": dst,
            "checkpoint":        meta.get("checkpoint", ""),
            "gold_state":        r["gold_state"],
            "gold_decision":     r["gold_decision"],
        }

        # GT 개수 입력 칸 (직접 작성)
        for obj in all_objects:
            row[f"gt_n_{obj}"] = ""

        # 탐지기별 × 타임스텝별 탐지 개수 + 노트
        # 순서: gdino [t0 counts | t0_note] [C1 counts | C1_note] [C2 counts | C2_note]
        #       molmo 동일
        for det in ("gdino", "molmo"):
            for frame in ("t0", "C1", "C2"):
                for obj in all_objects:
                    col = f"{det}_{frame}_{obj}"
                    row[col] = r.get(col, "")
                row[f"{det}_{frame}_note"] = ""     # 타임스텝별 노트 입력 칸

        # 자동 플래그
        row["gdino_target_flag"] = flag(g("gdino","t0",tgt), g("gdino","C1",tgt), g("gdino","C2",tgt))
        row["gdino_dest_flag"]   = flag(g("gdino","t0",dst), g("gdino","C1",dst), g("gdino","C2",dst))
        row["molmo_target_flag"] = flag(g("molmo","t0",tgt), g("molmo","C1",tgt), g("molmo","C2",tgt))
        row["molmo_dest_flag"]   = flag(g("molmo","t0",dst), g("molmo","C1",dst), g("molmo","C2",dst))

        # 피드백 빈칸
        row["issue_type"]        = ""
        row["affected_detector"] = ""
        row["severity"]          = ""
        row["notes"]             = ""

        data.append(row)

    df = pd.DataFrame(data)

    # ── XLSX 저장 + 서식 ───────────────────────────────────────────────────────
    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="feedback", index=False)
        ws = writer.sheets["feedback"]

        # 헤더 행 서식
        col_headers = list(df.columns)
        thin = Side(style="thin", color="AAAAAA")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for ci, col in enumerate(col_headers, start=1):
            cell = ws.cell(row=1, column=ci)
            cell.font      = Font(bold=True, size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border    = border

            # 섹션별 색상
            if col in ("id","scenario","task","objects",
                       "target_label","destination_label","checkpoint",
                       "gold_state","gold_decision"):
                cell.fill = hex_fill(HEADER_COLORS["meta"])
            elif col.startswith("gt_n_"):
                cell.fill = hex_fill(HEADER_COLORS["gt"])
            elif col.endswith("_note"):
                det = "gdino" if col.startswith("gdino") else "molmo"
                cell.fill = hex_fill(HEADER_COLORS[f"{det}_note"])
            elif col.startswith("gdino") and not col.endswith("flag"):
                cell.fill = hex_fill(HEADER_COLORS["gdino"])
            elif col.startswith("molmo") and not col.endswith("flag"):
                cell.fill = hex_fill(HEADER_COLORS["molmo"])
            elif col.endswith("flag"):
                cell.fill = hex_fill(HEADER_COLORS["flag"])
            elif col in ("issue_type","affected_detector","severity","notes"):
                cell.fill = hex_fill(HEADER_COLORS["feedback"])

        # 객체 탐지 컬럼: 객체별로 배경색 다르게 (헤더 + 데이터 행)
        for ci, col in enumerate(col_headers, start=1):
            if "_t0_" in col or "_C1_" in col or "_C2_" in col:
                parts = col.split("_", 2)
                if len(parts) == 3:
                    obj_name = parts[2]
                    fill_color = OBJ_COLORS.get(obj_name, DEFAULT_OBJ_COLOR)
                    # 데이터 행 배경 (연하게)
                    light = hex_fill(fill_color)
                    for ri in range(2, len(data) + 2):
                        ws.cell(row=ri, column=ci).fill = light

        # gold_decision 컬럼 텍스트 색상
        gd_col = col_headers.index("gold_decision") + 1
        decision_colors = {"CONTINUE": "1B7A3E", "ASK": "B45309", "STOP": "B91C1C"}
        for ri in range(2, len(data) + 2):
            cell = ws.cell(row=ri, column=gd_col)
            color = decision_colors.get(str(cell.value), "000000")
            cell.font = Font(bold=True, color=color)

        # 플래그 셀 강조 (비어있지 않은 경우 배경)
        flag_cols = [i+1 for i,c in enumerate(col_headers) if c.endswith("flag")]
        flag_fill = hex_fill("FFCCCC")
        for ci in flag_cols:
            for ri in range(2, len(data) + 2):
                cell = ws.cell(row=ri, column=ci)
                if cell.value:
                    cell.fill = flag_fill
                    cell.font = Font(bold=True, color="990000", size=9)

        # 열 너비 자동 조정
        col_widths = {
            "id": 16, "scenario": 12, "task": 32, "objects": 18, "n_objects": 8,
            "target_label": 12, "destination_label": 16,
            "checkpoint": 10, "gold_state": 10, "gold_decision": 12,
            "issue_type": 16, "affected_detector": 14, "severity": 10, "notes": 30,
        }
        for ci, col in enumerate(col_headers, start=1):
            w = col_widths.get(col, 11)
            ws.column_dimensions[get_column_letter(ci)].width = w

        # 헤더 행 높이 + 첫 행 고정
        ws.row_dimensions[1].height = 30
        ws.freeze_panes = "A2"

        # note 칸: 넉넉한 너비 + wrap_text
        note_cols = [i+1 for i, c in enumerate(col_headers) if c.endswith("_note")]
        note_border = Border(
            left=Side(style="medium", color="888888"),
            right=Side(style="medium", color="888888"),
            top=Side(style="thin",   color="AAAAAA"),
            bottom=Side(style="thin", color="AAAAAA"),
        )
        for ci in note_cols:
            col_letter = get_column_letter(ci)
            ws.column_dimensions[col_letter].width = 28
            det = "gdino" if col_headers[ci-1].startswith("gdino") else "molmo"
            fill = hex_fill(HEADER_COLORS[f"{det}_note"])
            for ri in range(2, len(data) + 2):
                cell = ws.cell(row=ri, column=ci)
                cell.fill      = fill
                cell.border    = note_border
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        # GT 입력 칸: 연한 갈색 배경 + 굵은 테두리로 "채워야 할 셀" 표시
        gt_cols = [i+1 for i, c in enumerate(col_headers) if c.startswith("gt_n_")]
        gt_border = Border(
            left=Side(style="medium", color="B8860B"),
            right=Side(style="medium", color="B8860B"),
            top=Side(style="thin", color="B8860B"),
            bottom=Side(style="thin", color="B8860B"),
        )
        gt_fill = hex_fill("FFF3CD")
        for ci in gt_cols:
            for ri in range(2, len(data) + 2):
                cell = ws.cell(row=ri, column=ci)
                cell.fill   = gt_fill
                cell.border = gt_border
                cell.alignment = Alignment(horizontal="center")

        # 자동 필터
        ws.auto_filter.ref = ws.dimensions

    print(f"저장 완료: {out_path}  ({len(data)}행, {len(col_headers)}열)")
    print(f"객체 종류: {all_objects}")

    # 플래그 통계
    flag_counts = collections.Counter()
    for r in rows:
        meta = manifest.get(r["id"], {})
        tgt  = meta.get("target_label", r["objects"][0] if r["objects"] else "")
        dst  = meta.get("destination_label", r["objects"][1] if len(r["objects"]) > 1 else "")
        def g(det, fr, obj): return r.get(f"{det}_{fr}_{obj}", 0)
        for det in ("gdino", "molmo"):
            for obj in (tgt, dst):
                fl = flag(g(det,"t0",obj), g(det,"C1",obj), g(det,"C2",obj))
                if fl:
                    flag_counts[f"{det}_{fl}"] += 1

    print("\n자동 플래그 통계:")
    for k, v in sorted(flag_counts.items(), key=lambda x: -x[1]):
        print(f"  {k:<35} {v}건")


if __name__ == "__main__":
    main()
